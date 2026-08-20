"""
core/comparaison.py
====================

Étude comparative QR Code entre deux dates. Rapproche les agents des deux
extractions par numéro de téléphone (pos_msisdn), exactement comme dans
l'appli HTML. Reprend le format à 3 onglets déjà validé (calé sur l'exemple
fourni) : Résumé comparatif (formules Excel vivantes), Répartition par
catégorie, Mouvements détaillés (regroupés par catégorie x statut, avec
uniquement les sections où il y a eu du mouvement).
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.qr_code import STATUTS, STATUT_PREFIX, STATUT_SHORT

FONT_NAME = "Arial"
HEADER_FILL = "1F4E78"
BAND_FILL = "F2F6FA"
TOTAL_BG = "EAF1F8"
GREY = "666666"
RED = "B3352C"
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COUNT_ROW_ORDER = [
    ("Total agents (base)", None),
    ("Sans QR Code", "Sans QR Code"),
    ("QR reçu, non utilisé (+30j)", "QR non utilisé (+30j)"),
    ("Risque d'inactivité (20-29j)", "Risque inactivité"),
    ("QR actifs (utilisés < 20j)", "Actif"),
]


def _key_msisdn(v) -> str:
    """Normalise un numéro de téléphone en clé de rapprochement (gère les '.0' numériques parasites)."""
    if pd.isna(v):
        return ""
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def _style_header(ws, row, ncols, size=10):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT_NAME, size=size, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_label(cell, bold=False):
    cell.font = Font(name=FONT_NAME, size=10, bold=bold)
    cell.border = BORDER


def _style_value(cell, is_pct=False, bold=False, center=True, numfmt=None, color=None, fill=None):
    cell.font = Font(name=FONT_NAME, size=10, bold=bold, color=color)
    cell.border = BORDER
    if center:
        cell.alignment = Alignment(horizontal="center")
    if is_pct:
        cell.number_format = "0.00%"
    if numfmt:
        cell.number_format = numfmt
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)


def _counts_for(df: pd.DataFrame) -> dict:
    c = {"Total agents (base)": len(df)}
    for statut in STATUTS:
        c[statut] = int((df["statut"] == statut).sum())
    return c


def _write_counts_table(ws, start_row, headers, counts_a, counts_b, with_pct):
    r = start_row
    for i, h in enumerate(headers, start=1):
        ws.cell(row=r, column=i, value=h)
    _style_header(ws, r, len(headers), size=10 if with_pct else 9)
    r += 1
    first_data_row = r
    for label, key in COUNT_ROW_ORDER:
        k = key or "Total agents (base)"
        va, vb = counts_a[k], counts_b[k]
        _style_label(ws.cell(row=r, column=1, value=label), bold=(label == "Total agents (base)"))
        _style_value(ws.cell(row=r, column=2, value=va))
        _style_value(ws.cell(row=r, column=3, value=vb))
        evol = ws.cell(row=r, column=4, value=f"=C{r}-B{r}")
        _style_value(evol, bold=True, numfmt="+#,##0;-#,##0")
        if with_pct:
            pct = ws.cell(row=r, column=5, value=f'=IF(B{r}=0,"",(C{r}-B{r})/B{r})')
            _style_value(pct, numfmt="+0.00%;-0.00%")
        r += 1
    return r, first_data_row


def build_comparative_workbook(df_a: pd.DataFrame, df_b: pd.DataFrame,
                                label_a: str, label_b: str, source_label: str = "ALBARKA") -> Workbook:
    """
    df_a, df_b : DataFrames déjà classifiés (colonnes 'statut', 'segment_group',
                 'pos_msisdn', 'pos_name', 'dsm_name', 'region', 'territory', 'site_name', ...)
                 -- c'est-à-dire le résultat de qr_code.classify(...).
    label_a, label_b : libellés d'affichage des deux dates (ex. "26/07/2026").
    """
    segments = sorted(set(df_a["segment_group"].dropna()) | set(df_b["segment_group"].dropna()))
    wb = Workbook()
    wb.remove(wb.active)

    # ================= Feuille 1 : Résumé comparatif =================
    ws = wb.create_sheet("Résumé comparatif")
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{source_label} — ÉTUDE COMPARATIVE"
    ws["A1"].font = Font(name=FONT_NAME, size=15, bold=True, color=HEADER_FILL)
    ws["A2"] = f"Comparaison {label_a} vs {label_b}"
    ws["A2"].font = Font(name=FONT_NAME, size=10, italic=True, color=GREY)

    counts_a, counts_b = _counts_for(df_a), _counts_for(df_b)

    r = 4
    ws.cell(row=r, column=1, value="RÉPARTITION PAR STATUT — GLOBAL").font = \
        Font(name=FONT_NAME, size=11, bold=True, color=HEADER_FILL)
    r += 1
    r, first_row = _write_counts_table(ws, r, ["Statut", label_a, label_b, "Évolution", "Évolution %"],
                                        counts_a, counts_b, with_pct=True)
    r += 1
    row_total, row_sans, row_nonutil, row_risque, row_actif = first_row, first_row + 1, first_row + 2, first_row + 3, first_row + 4

    ws.cell(row=r, column=1, value="INDICATEURS CLÉS GLOBAUX").font = \
        Font(name=FONT_NAME, size=11, bold=True, color=HEADER_FILL)
    r += 1
    for i, h in enumerate(["Indicateur", label_a, label_b, "Évolution (pts)"], start=1):
        ws.cell(row=r, column=i, value=h)
    _style_header(ws, r, 4)
    r += 1
    kpis = [
        ("Taux de déploiement QR Code (QR distribués / total agents)",
         f"=1-B{row_sans}/B{row_total}", f"=1-C{row_sans}/C{row_total}"),
        ("Taux d'utilisation parmi les QR déployés (actifs / QR déployés)",
         f"=B{row_actif}/(B{row_total}-B{row_sans})", f"=C{row_actif}/(C{row_total}-C{row_sans})"),
        ("Taux de QR déployés mais non utilisés (+30j) / QR déployés",
         f"=B{row_nonutil}/(B{row_total}-B{row_sans})", f"=C{row_nonutil}/(C{row_total}-C{row_sans})"),
        ("Taux de risque d'inactivité (/ total agents)",
         f"=B{row_risque}/B{row_total}", f"=C{row_risque}/C{row_total}"),
        ("Part des agents sans QR Code sur le total",
         f"=B{row_sans}/B{row_total}", f"=C{row_sans}/C{row_total}"),
    ]
    for label, fa, fb in kpis:
        _style_label(ws.cell(row=r, column=1, value=label))
        ca = ws.cell(row=r, column=2, value=fa)
        _style_value(ca, is_pct=True, bold=True, color=HEADER_FILL, fill=TOTAL_BG)
        cb = ws.cell(row=r, column=3, value=fb)
        _style_value(cb, is_pct=True, bold=True, color=HEADER_FILL, fill=TOTAL_BG)
        cd = ws.cell(row=r, column=4, value=f"=(C{r}-B{r})*100")
        _style_value(cd, bold=True, numfmt="+0.00;-0.00")
        r += 1
    r += 1
    ws.cell(row=r, column=1, value='Note : le détail par catégorie figure dans l\'onglet "Répartition par '
                                    'catégorie". Les mouvements agent par agent figurent dans l\'onglet '
                                    '"Mouvements détaillés".')
    ws.cell(row=r, column=1).font = Font(name=FONT_NAME, size=9, italic=True, color="888888")

    ws.column_dimensions["A"].width = 58
    for c in "BCDE":
        ws.column_dimensions[c].width = 14

    # ================= Feuille 2 : Répartition par catégorie =================
    ws2 = wb.create_sheet("Répartition par catégorie")
    ws2.sheet_view.showGridLines = False
    ws2["A1"] = "RÉPARTITION PAR CATÉGORIE"
    ws2["A1"].font = Font(name=FONT_NAME, size=13, bold=True, color=HEADER_FILL)
    ws2["A2"] = f"Comparaison {label_a} vs {label_b}"
    ws2["A2"].font = Font(name=FONT_NAME, size=10, italic=True, color=GREY)

    r = 4
    for seg in segments:
        ws2.cell(row=r, column=1, value=seg).font = Font(name=FONT_NAME, size=11, bold=True, color=HEADER_FILL)
        r += 1
        sub_a = df_a[df_a["segment_group"] == seg]
        sub_b = df_b[df_b["segment_group"] == seg]
        r, _ = _write_counts_table(ws2, r, ["Statut", label_a, label_b, "Évolution"],
                                    _counts_for(sub_a), _counts_for(sub_b), with_pct=False)
        r += 1

    ws2.column_dimensions["A"].width = 30
    for c in "BCD":
        ws2.column_dimensions[c].width = 14

    # ================= Feuille 3 : Mouvements détaillés =================
    ws3 = wb.create_sheet("Mouvements détaillés")
    ws3.sheet_view.showGridLines = False
    ws3["A1"] = "MOUVEMENTS DÉTAILLÉS PAR AGENT"
    ws3["A1"].font = Font(name=FONT_NAME, size=13, bold=True, color=HEADER_FILL)
    ws3["A2"] = f"Entrées et sorties de chaque liste entre le {label_a} et le {label_b}"
    ws3["A2"].font = Font(name=FONT_NAME, size=10, italic=True, color=GREY)

    df_a = df_a.copy()
    df_b = df_b.copy()
    df_a["_key"] = df_a["pos_msisdn"].apply(_key_msisdn)
    df_b["_key"] = df_b["pos_msisdn"].apply(_key_msisdn)

    table_headers = ["Mouvement", "Agent (POS)", "Téléphone", "DSM", "Région", "Territoire", "Site"]
    col_widths = [12, 32, 13, 12, 10, 18, 20]
    detail_order = ["Sans QR Code", "QR non utilisé (+30j)", "Risque inactivité"]

    r = 4
    for statut in detail_order:
        for seg in segments:
            in_b = df_b[(df_b["segment_group"] == seg) & (df_b["statut"] == statut)]
            in_a = df_a[(df_a["segment_group"] == seg) & (df_a["statut"] == statut)]
            keys_a_same_statut = set(in_a["_key"]) - {""}
            keys_b_same_statut = set(in_b["_key"]) - {""}

            entrees = in_b[~in_b["_key"].isin(keys_a_same_statut) & (in_b["_key"] != "")]
            sorties = in_a[~in_a["_key"].isin(keys_b_same_statut) & (in_a["_key"] != "")]

            if len(entrees) == 0 and len(sorties) == 0:
                continue

            sheet_key = f"{STATUT_PREFIX[statut]}.{seg}-{STATUT_SHORT[statut]}"
            cell = ws3.cell(row=r, column=1,
                             value=f"{sheet_key}  (entrées : {len(entrees)}  |  sorties : {len(sorties)})")
            cell.font = Font(name=FONT_NAME, size=10.5, bold=True, color=HEADER_FILL)
            ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(table_headers))
            r += 1

            for i, h in enumerate(table_headers, start=1):
                ws3.cell(row=r, column=i, value=h)
            _style_header(ws3, r, len(table_headers), size=9)
            r += 1

            for mouvement, sub, color in [("Entrée", entrees, HEADER_FILL), ("Sortie", sorties, RED)]:
                for _, agent in sub.iterrows():
                    vals = [mouvement, agent.get("pos_name"), agent.get("pos_msisdn"), agent.get("dsm_name"),
                            agent.get("region"), agent.get("territory"), agent.get("site_name")]
                    for i, v in enumerate(vals, start=1):
                        cell = ws3.cell(row=r, column=i, value=v if pd.notna(v) else "")
                        cell.font = Font(name=FONT_NAME, size=10, bold=(i == 1),
                                          color=color if i == 1 else None)
                        cell.border = BORDER
                    r += 1
            r += 1

    for i, w in enumerate(col_widths, start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    return wb


# ---------------------------------------------------------------------------
# Auto-test : compare deux vrais fichiers QR Code par numéro de téléphone.
# python3 -m core.comparaison <fichier_A> <date_A> <fichier_B> <date_B>
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import sys
    from core.qr_code import read_qr_file, classify

    if len(sys.argv) < 5:
        print("Usage : python3 -m core.comparaison <fichier_A> <date_A AAAA-MM-JJ> <fichier_B> <date_B AAAA-MM-JJ>")
        sys.exit(1)

    fichier_a, date_a, fichier_b, date_b = sys.argv[1:5]

    print(f"Lecture et classification de {fichier_a} (référence {date_a}) ...")
    df_a = classify(read_qr_file(fichier_a), date_a)
    print(f"Lecture et classification de {fichier_b} (référence {date_b}) ...")
    df_b = classify(read_qr_file(fichier_b), date_b)

    print(f"\n{date_a} : {len(df_a)} agents — {df_a['statut'].value_counts().to_dict()}")
    print(f"{date_b} : {len(df_b)} agents — {df_b['statut'].value_counts().to_dict()}")

    label_a = pd.Timestamp(date_a).strftime("%d/%m/%Y")
    label_b = pd.Timestamp(date_b).strftime("%d/%m/%Y")

    wb = build_comparative_workbook(df_a, df_b, label_a, label_b, source_label="ALBARKA")
    out_path = f"/tmp/test_comparatif_{date_a}_vs_{date_b}.xlsx"
    wb.save(out_path)
    print(f"\nRapport généré : {out_path}")
    print(f"Onglets : {wb.sheetnames}")
