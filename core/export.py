"""
core/export.py
===============

Utilitaires d'export Excel réutilisables dans toutes les pages Streamlit.

Fonction principale :
  export_df_to_excel(sheets, titre, source_label) -> bytes

  `sheets` est un dict ordonné {nom_onglet: pd.DataFrame}.
  Retourne les bytes du classeur Excel (.xlsx), prêts à être passés à
  st.download_button(..., data=<bytes>, ...).

Style appliqué :
  - En-tête de colonne : fond bleu foncé (1F4E78), texte blanc, gras
  - Lignes alternées : fond bleu très clair (F2F6FA)
  - Ligne de titre (A1) avec le libellé du rapport
  - Colonnes auto-dimensionnées (longueur max du contenu, cappée à 60)
  - Thème cohérent avec le reste de l'application (même palette)

Usage dans une page Streamlit :
    from core.export import export_df_to_excel

    data = export_df_to_excel({"Classement": df_classement, "Alertes": df_alertes},
                               titre="Cash Flow — Juillet 2026")
    st.download_button(" Exporter Excel", data=data,
                       file_name="cashflow_2026-07.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
"""

import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constantes de style (cohérentes avec qr_code.py / comparaison.py)
# ---------------------------------------------------------------------------
FONT_NAME   = "Arial"
HEADER_FILL = "1F4E78"
BAND_FILL   = "F2F6FA"
GREY        = "666666"
THIN        = Side(style="thin", color="D9D9D9")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header_cell(cell):
    cell.font      = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
    cell.fill      = PatternFill("solid", fgColor=HEADER_FILL)
    cell.border    = BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_data_cell(cell, band: bool = False):
    cell.font   = Font(name=FONT_NAME, size=10)
    cell.border = BORDER
    if band:
        cell.fill = PatternFill("solid", fgColor=BAND_FILL)


def _auto_width(ws, col_idx: int, values: list, min_w: int = 8, max_w: int = 60):
    """Ajuste la largeur d'une colonne d'après le contenu le plus long."""
    max_len = max((len(str(v)) for v in values if v is not None), default=min_w)
    ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, min_w), max_w)


def _sanitize_sheet_name(name: str) -> str:
    """
    Supprime les caractères interdits par openpyxl dans les noms d'onglets Excel
    (/ \\ ? * [ ] : ) et tronque à 31 caractères.
    """
    for ch in r'/\?*[]:\\':
        name = name.replace(ch, "-")
    return name[:31]


def _write_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame,
                 titre: str, source_label: str):
    """Écrit un DataFrame dans un onglet avec mise en forme complète."""
    ws = wb.create_sheet(_sanitize_sheet_name(sheet_name))
    ws.sheet_view.showGridLines = False

    # Ligne titre
    ws["A1"] = titre
    ws["A1"].font = Font(name=FONT_NAME, size=13, bold=True, color=HEADER_FILL)
    ws["A2"] = f"Exporté le {datetime.now().strftime('%d/%m/%Y à %H:%M')} — {source_label}"
    ws["A2"].font = Font(name=FONT_NAME, size=9, italic=True, color=GREY)

    if df.empty:
        ws["A4"] = "Aucune donnée à afficher."
        ws["A4"].font = Font(name=FONT_NAME, size=10, italic=True, color=GREY)
        return

    header_row = 4
    cols = list(df.columns)

    # En-têtes
    for j, col in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=j, value=str(col))
        _style_header_cell(cell)

    # Données
    for i, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        band = (i % 2 == 0)
        for j, col in enumerate(cols, start=1):
            val = row[col]
            # Convertit les types pandas non sérialisables
            if pd.isna(val) if not isinstance(val, (list, dict)) else False:
                val = ""
            elif hasattr(val, "item"):
                val = val.item()
            cell = ws.cell(row=i, column=j, value=val)
            _style_data_cell(cell, band=band)
            cell.alignment = Alignment(horizontal="left")

    # Largeurs automatiques
    for j, col in enumerate(cols, start=1):
        col_values = [col] + [str(row[col]) for _, row in df.iterrows()]
        _auto_width(ws, j, col_values)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def export_df_to_excel(sheets: dict, titre: str,
                       source_label: str = "ALBARKA") -> bytes:
    """
    Génère un classeur Excel multi-onglets à partir d'un dict
    {nom_onglet: pd.DataFrame}.

    Paramètres :
      sheets       : dict ordonné {str: pd.DataFrame}
      titre        : titre affiché en A1 de chaque onglet
      source_label : mention de la source (ex. "ALBARKA — Cash Flow")

    Retourne les bytes du fichier .xlsx.
    """
    wb = Workbook()
    wb.remove(wb.active)  # supprime la feuille vide par défaut

    for sheet_name, df in sheets.items():
        if not isinstance(df, pd.DataFrame):
            df = pd.DataFrame(df)
        _write_sheet(wb, sheet_name, df, titre=titre, source_label=source_label)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
