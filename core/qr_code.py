"""
core/qr_code.py
================

Nettoyage, classification et génération du rapport pour le suivi QR Code des
agents ALBARKA. Reprend exactement la logique déjà validée (Excel/Python puis
JavaScript) : même règle de classification, même structure de rapport
(Résumé + 9 onglets détaillés par catégorie x statut), même mise en forme
bleu/blanc.
"""

import gzip
import io
from pathlib import Path
from datetime import datetime, date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constantes de style (thème bleu / blanc)
# ---------------------------------------------------------------------------
FONT_NAME = "Arial"
HEADER_FILL = "1F4E78"
BAND_FILL = "F2F6FA"
TOTAL_BG = "EAF1F8"
GREY = "666666"
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUTS = ["Sans QR Code", "QR non utilisé (+30j)", "Risque inactivité", "Actif"]
STATUT_PREFIX = {"Sans QR Code": "1", "QR non utilisé (+30j)": "2", "Risque inactivité": "3"}
STATUT_SHORT = {"Sans QR Code": "Sans QR Code", "QR non utilisé (+30j)": "QR non utilisé",
                "Risque inactivité": "Risque inactivité"}
STATUT_DESC = {
    "Sans QR Code": "Agents n'ayant pas encore reçu de QR Code",
    "QR non utilisé (+30j)": "Agents ayant un QR Code mais ne l'utilisant pas (inactifs +30 jours)",
    "Risque inactivité": "Agents à risque de devenir inactifs (dernière utilisation 20-29 jours)",
}
COLS_COMMON = ["dsm_name", "region", "territory", "town", "quartier", "site_name", "pos_name", "pos_msisdn"]
RENAME_MAP = {
    "dsm_name": "DSM", "region": "Région", "territory": "Territoire", "town": "Ville",
    "quartier": "Quartier", "site_name": "Site", "pos_name": "Agent (POS)", "pos_msisdn": "Téléphone",
    "last_qr_co_date": "Dernière utilisation QR", "days_since_last_use": "Jours sans utilisation",
    "priorite": "Priorité",
}


# ---------------------------------------------------------------------------
# Lecture du fichier source (gère .xlsx direct ou .xlsx compressé en .gz)
# ---------------------------------------------------------------------------
def read_qr_file(source) -> pd.DataFrame:
    """
    Lit le fichier QR Code source. `source` peut être :
    - un chemin de fichier (str ou Path) vers un .xlsx ou un .gz
    - un objet bytes (ex. fichier uploadé via Streamlit)
    Détecte automatiquement s'il s'agit d'une archive gzip (signature 0x1f 0x8b).
    """
    if isinstance(source, (str, Path)):
        raw = Path(source).read_bytes()
    else:
        raw = source.read() if hasattr(source, "read") else bytes(source)

    if raw[:2] == b"\x1f\x8b":
        raw = gzip.decompress(raw)

    return pd.read_excel(io.BytesIO(raw))


# ---------------------------------------------------------------------------
# Classification
# ---------------------------------------------------------------------------
def classify(df: pd.DataFrame, ref_date) -> pd.DataFrame:
    """
    Ajoute au DataFrame les colonnes 'days_since_last_use', 'statut' et
    'priorite', selon la règle métier déjà validée :
      - active_deployed vide            -> Sans QR Code
      - active_30 = 0                   -> QR non utilisé (+30j)          (priorité 2)
      - jours sans usage >= 20          -> Risque inactivité              (priorité 1)
      - sinon                           -> Actif
    """
    df = df.copy()
    if isinstance(ref_date, date) and not isinstance(ref_date, datetime):
        ref_date = datetime(ref_date.year, ref_date.month, ref_date.day)
    ref_date = pd.Timestamp(ref_date)

    df["last_qr_co_date"] = pd.to_datetime(df["last_qr_co_date"], errors="coerce")
    df["days_since_last_use"] = (ref_date - df["last_qr_co_date"]).dt.days

    def _statut(row):
        if pd.isna(row["active_deployed"]):
            return "Sans QR Code"
        if row["active_30"] == 0:
            return "QR non utilisé (+30j)"
        if row["days_since_last_use"] >= 20:
            return "Risque inactivité"
        return "Actif"

    df["statut"] = df.apply(_statut, axis=1)
    df["priorite"] = df["statut"].map({"Risque inactivité": 1, "QR non utilisé (+30j)": 2})
    return df


# ---------------------------------------------------------------------------
# Styles (aides internes pour la génération du classeur)
# ---------------------------------------------------------------------------
def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_label(cell, bold=False):
    cell.font = Font(name=FONT_NAME, size=10, bold=bold)
    cell.border = BORDER


def _style_value(cell, is_pct=False, bold=False, color=None, fill=None):
    cell.font = Font(name=FONT_NAME, size=10, bold=bold, color=color)
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="center")
    if is_pct:
        cell.number_format = "0.0%"
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)


def _write_detail_sheet(wb, sheet_name, title, subtitle, df_data, with_dates):
    ws = wb.create_sheet(sheet_name[:31])
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT_NAME, size=13, bold=True, color=HEADER_FILL)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name=FONT_NAME, size=10, italic=True, color=GREY)

    cols = COLS_COMMON + (["last_qr_co_date", "days_since_last_use"] if with_dates else [])
    headers = [RENAME_MAP.get(c, c) for c in cols]
    start_row = 4
    for j, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=j, value=h)
    _style_header(ws, start_row, len(headers))

    for i, (_, row) in enumerate(df_data.iterrows(), start=start_row + 1):
        for j, col in enumerate(cols, start=1):
            val = row[col]
            if col == "last_qr_co_date" and pd.notna(val):
                val = val.strftime("%d/%m/%Y")
            cell = ws.cell(row=i, column=j, value=val if pd.notna(val) else "")
            cell.font = Font(name=FONT_NAME, size=10)
            cell.border = BORDER
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=BAND_FILL)

    widths = {"DSM": 12, "Région": 10, "Territoire": 16, "Ville": 10, "Quartier": 16, "Site": 20,
              "Agent (POS)": 32, "Téléphone": 13, "Dernière utilisation QR": 14, "Jours sans utilisation": 12}
    for j, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(h, 14)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


# ---------------------------------------------------------------------------
# Construction du classeur complet
# ---------------------------------------------------------------------------
def build_report_workbook(df_classified: pd.DataFrame, ref_date, source_label: str = "ALBARKA") -> Workbook:
    """
    Construit le classeur complet : onglet Résumé (comptes, pourcentages,
    indicateurs clés, formules Excel vivantes) + un onglet détaillé par
    combinaison catégorie x statut (Sans QR Code / QR non utilisé / Risque
    d'inactivité), pour chacune des catégories présentes dans les données.
    """
    if isinstance(ref_date, (date,)) and not isinstance(ref_date, datetime):
        ref_date_dt = ref_date
    else:
        ref_date_dt = pd.Timestamp(ref_date).date()

    segments = sorted(df_classified["segment_group"].dropna().unique().tolist())
    wb = Workbook()
    wb.remove(wb.active)

    # ---- Résumé ----
    ws = wb.create_sheet("Résumé")
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{source_label} — Suivi QR Code Agents"
    ws["A1"].font = Font(name=FONT_NAME, size=16, bold=True, color=HEADER_FILL)
    ws["A2"] = f"Date de référence : {ref_date_dt.strftime('%d/%m/%Y')}"
    ws["A2"].font = Font(name=FONT_NAME, size=10, italic=True, color=GREY)

    r = 4
    ws.cell(row=r, column=1, value="RÉPARTITION PAR STATUT").font = Font(name=FONT_NAME, size=11, bold=True, color=HEADER_FILL)
    r += 1
    headers = ["Statut"] + segments + ["Total"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    _style_header(ws, r, len(headers))
    r += 1

    seg_totals = df_classified["segment_group"].value_counts()
    counts = {"Total agents (base)": [int(seg_totals.get(s, 0)) for s in segments]}
    for statut in STATUTS[:3]:  # Sans QR / Non utilisé / Risque (Actif = calculé par différence)
        sub = df_classified[df_classified["statut"] == statut]
        counts[statut] = [int((sub["segment_group"] == s).sum()) for s in segments]
    counts["Actif"] = [counts["Total agents (base)"][i] - counts["Sans QR Code"][i]
                        - counts["QR non utilisé (+30j)"][i] - counts["Risque inactivité"][i]
                        for i in range(len(segments))]

    row_labels = [
        ("Total agents (base)", "Total agents (base)"),
        ("Sans QR Code", "Sans QR Code"),
        ("QR reçu, non utilisé (+30j)", "QR non utilisé (+30j)"),
        ("Risque d'inactivité (20-29j)", "Risque inactivité"),
        ("QR actifs (utilisés < 20j)", "Actif"),
    ]
    row_index = {}
    for label, key in row_labels:
        _style_label(ws.cell(row=r, column=1, value=label), bold=(key == "Total agents (base)"))
        for i, seg in enumerate(segments):
            ws.cell(row=r, column=2 + i, value=counts[key][i])
            _style_value(ws.cell(row=r, column=2 + i))
        total_col = 2 + len(segments)
        ws.cell(row=r, column=total_col, value=f"=SUM(B{r}:{get_column_letter(total_col - 1)}{r})")
        _style_value(ws.cell(row=r, column=total_col), bold=True)
        row_index[key] = r
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="RÉPARTITION EN % (par rapport au total d'agents)").font = \
        Font(name=FONT_NAME, size=11, bold=True, color=HEADER_FILL)
    r += 1
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    _style_header(ws, r, len(headers))
    r += 1
    pct_labels = ["Sans QR Code", "QR non utilisé (+30j)", "Risque inactivité", "Actif"]
    pct_display = ["Sans QR Code", "QR reçu, non utilisé (+30j)", "Risque d'inactivité (20-29j)", "QR actifs (utilisés < 20j)"]
    for disp, key in zip(pct_display, pct_labels):
        _style_label(ws.cell(row=r, column=1, value=f"% {disp}"))
        src_row = row_index[key]
        tot_row = row_index["Total agents (base)"]
        for c in range(2, 2 + len(segments) + 1):
            letter = get_column_letter(c)
            cell = ws.cell(row=r, column=c, value=f"={letter}{src_row}/{letter}{tot_row}")
            _style_value(cell, is_pct=True)
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="INDICATEURS CLÉS GLOBAUX").font = Font(name=FONT_NAME, size=11, bold=True, color=HEADER_FILL)
    r += 1
    total_col_letter = get_column_letter(2 + len(segments))
    tot_row = row_index["Total agents (base)"]
    sans_row = row_index["Sans QR Code"]
    nonutil_row = row_index["QR non utilisé (+30j)"]
    risque_row = row_index["Risque inactivité"]
    actif_row = row_index["Actif"]
    kpis = [
        ("Taux de déploiement QR Code (QR distribués / total agents)",
         f"=({total_col_letter}{tot_row}-{total_col_letter}{sans_row})/{total_col_letter}{tot_row}"),
        ("Taux d'utilisation parmi les QR déployés (actifs / QR déployés)",
         f"={total_col_letter}{actif_row}/({total_col_letter}{tot_row}-{total_col_letter}{sans_row})"),
        ("Taux de QR déployés mais non utilisés (+30j) / QR déployés",
         f"={total_col_letter}{nonutil_row}/({total_col_letter}{tot_row}-{total_col_letter}{sans_row})"),
        ("Taux de risque d'inactivité (/ total agents)",
         f"={total_col_letter}{risque_row}/{total_col_letter}{tot_row}"),
        ("Part des agents sans QR Code sur le total",
         f"={total_col_letter}{sans_row}/{total_col_letter}{tot_row}"),
    ]
    for label, formula in kpis:
        _style_label(ws.cell(row=r, column=1, value=label))
        cell = ws.cell(row=r, column=2, value=formula)
        _style_value(cell, is_pct=True, bold=True, color=HEADER_FILL, fill=TOTAL_BG)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value=f"Note : chaque catégorie ({' / '.join(segments)}) dispose de son "
                                    f"propre onglet pour chaque statut.")
    ws.cell(row=r, column=1).font = Font(name=FONT_NAME, size=9, italic=True, color="888888")

    ws.column_dimensions["A"].width = 55
    for c in range(2, 2 + len(segments) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13

    # ---- Onglets détaillés ----
    for statut in STATUTS[:3]:
        for seg in segments:
            sub = df_classified[(df_classified["statut"] == statut) & (df_classified["segment_group"] == seg)].copy()
            if statut in ("Risque inactivité", "QR non utilisé (+30j)"):
                sub = sub.sort_values("days_since_last_use", ascending=False)
            else:
                sub = sub.sort_values(["dsm_name", "pos_name"])
            sheet_name = f"{STATUT_PREFIX[statut]}.{seg}-{STATUT_SHORT[statut]}"
            title = f"{STATUT_SHORT[statut]} — {seg}"
            subtitle = f"{len(sub)} agent(s) — {STATUT_DESC[statut]}"
            _write_detail_sheet(wb, sheet_name, title, subtitle, sub, with_dates=(statut != "Sans QR Code"))

    return wb


# ---------------------------------------------------------------------------
# Auto-test : lance ce fichier directement (python3 core/qr_code.py) pour
# vérifier le traitement complet sur un vrai fichier, sans dépendre de Streamlit.
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import sys

    chemin_test = sys.argv[1] if len(sys.argv) > 1 else None
    if not chemin_test:
        print("Usage : python3 core/qr_code.py <chemin_vers_fichier_qr.xlsx_ou_.gz> [date_reference AAAA-MM-JJ]")
        sys.exit(1)

    date_ref = sys.argv[2] if len(sys.argv) > 2 else None

    print(f"Lecture de {chemin_test} ...")
    df = read_qr_file(chemin_test)
    print(f"{len(df)} lignes lues.")

    if not date_ref:
        date_ref = datetime.now().strftime("%Y-%m-%d")
        print(f"Aucune date de référence fournie, utilisation de la date du jour : {date_ref}")

    df_classified = classify(df, date_ref)
    print("\nRépartition par statut :")
    print(df_classified["statut"].value_counts())

    wb = build_report_workbook(df_classified, date_ref, source_label="ALBARKA")
    out_path = f"/tmp/test_qr_report_{date_ref}.xlsx"
    wb.save(out_path)
    print(f"\nRapport généré : {out_path}")
    print(f"Onglets : {wb.sheetnames}")
