"""
core/transactions.py — v2
==========================

Nettoyage des fichiers CSV de transactions Mobile Money et génération
du classeur Excel (3 onglets).

Nouveautés v2 :
  - Colonnes From msisdn / To msisdn conservées (nécessaires pour le
    rapprochement portefeuille et le stockage clients_servis)
  - Points touchés = nb de lignes après nettoyage (comptage brut, pas de
    déduplication — conforme au CDC)
  - extract_clients_servis() : extrait les contreparties par commercial
    (alias) et par jour → alimentation de la table clients_servis
  - extract_appro_from_tcd() : lit le classeur généré (onglets TCD) et
    extrait l'appro/destockage depuis les lignes de l'alias du commercial

Règle d'exclusion : ALBARKA GN SARL et ALBARKA GN SARL 5 sont toujours
exclus des calculs de contrepartie et des TCD.
"""

from pathlib import Path
from io import BytesIO

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME     = "Arial"
HEADER_FILL   = "1F4E78"
SUBHEADER_FILL = "2E75B6"
BAND_FILL     = "F2F6FA"
TOTAL_BG      = "EAF1F8"
GREY          = "666666"
THIN          = Side(style="thin", color="D9D9D9")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

EXCLUDED_NAMES = {"ALBARKA GN SARL", "ALBARKA GN SARL 5"}

# Colonnes obligatoires dans le CSV source
_COLS_REQUIRED = ["Date", "Type", "From name", "To name", "Amount"]
# Colonnes téléphone (optionnelles — présentes dans certains exports MTN)
_COLS_MSISDN   = ["From msisdn", "To msisdn"]


# ---------------------------------------------------------------------------
# Nettoyage
# ---------------------------------------------------------------------------

def clean_transactions(source) -> pd.DataFrame:
    """
    Lit et nettoie un fichier CSV de transactions Mobile Money.

    Colonnes conservées :
      - Date (réduite au jour)
      - Type
      - From name, To name
      - Amount
      - From msisdn, To msisdn  (si présentes dans le fichier source)

    Filtres :
      - Type = Transfer uniquement
      - Exclusion ALBARKA GN SARL / ALBARKA GN SARL 5

    `source` peut être un chemin (str/Path), un objet fichier ou des bytes.
    """
    if isinstance(source, (str, Path)):
        df = pd.read_csv(source)
    elif isinstance(source, bytes):
        df = pd.read_csv(BytesIO(source))
    else:
        df = pd.read_csv(source)

    df.columns = df.columns.str.strip()

    # Filtre Status = Successful si présent
    if "Status" in df.columns:
        df = df[df["Status"].str.strip() == "Successful"]

    # Vérifier les colonnes obligatoires
    missing = [c for c in _COLS_REQUIRED if c not in df.columns]
    if missing:
        raise ValueError(f"Colonnes manquantes dans le CSV : {missing}")

    # Colonnes à conserver (obligatoires + msisdn si présents)
    msisdn_present = [c for c in _COLS_MSISDN if c in df.columns]
    cols_keep = _COLS_REQUIRED + msisdn_present
    df = df[cols_keep].copy()

    # Nettoyage Amount
    if df["Amount"].dtype == object:
        df["Amount"] = (
            df["Amount"].astype(str)
            .str.replace(r"[\s,]", "", regex=True)
            .astype(float)
        )
    else:
        df["Amount"] = df["Amount"].astype(float)

    df["Date"] = pd.to_datetime(df["Date"]).dt.date
    df = df[df["Type"].str.strip() == "Transfer"]
    df = df[~df["To name"].isin(EXCLUDED_NAMES)]
    df = df[~df["From name"].isin(EXCLUDED_NAMES)]

    # Normaliser les MSISDN si présents
    for col in msisdn_present:
        df[col] = df[col].apply(_normalise_msisdn)

    return df.reset_index(drop=True)


def _normalise_msisdn(val) -> str:
    """
    Normalise un numéro MSISDN : supprime espaces/tirets, préfixe +237/00237,
    suffixe .0 parasite. Retourne chaîne vide si invalide.
    """
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    s = str(val).strip().replace(" ", "").replace("-", "").replace(".", "")
    if s.endswith("0") and s[:-1].isdigit() and len(s) > 9:
        # ex. "6990000000" issu de "699000000.0"
        # on ne tronque que si c'était du float avec .0
        raw = str(val).strip()
        if raw.endswith(".0"):
            s = raw[:-2].replace(" ", "").replace("-", "")
    for prefixe in ("+237", "00237"):
        if s.startswith(prefixe):
            s = s[len(prefixe):]
    return s


# ---------------------------------------------------------------------------
# Points touchés
# ---------------------------------------------------------------------------

def compute_points_touches(df: pd.DataFrame) -> dict:
    """
    Calcule les points touchés par jour et sur la période complète.
    Points touchés = nb de lignes (transactions) après nettoyage — pas de
    déduplication, conforme au CDC.

    Retourne :
    {
      "total": int,
      "par_jour": {"AAAA-MM-JJ": int, ...},
      "moyenne_par_jour": float,
    }
    """
    par_jour = {}
    for d, grp in df.groupby("Date"):
        par_jour[str(d)] = len(grp)

    total = len(df)
    nb_jours = len(par_jour)
    moyenne = round(total / nb_jours, 2) if nb_jours else 0.0

    return {
        "total":           total,
        "par_jour":        par_jour,
        "moyenne_par_jour": moyenne,
    }


# ---------------------------------------------------------------------------
# Extraction clients servis
# ---------------------------------------------------------------------------

def extract_clients_servis(df: pd.DataFrame, alias: str) -> list[dict]:
    """
    Extrait les contreparties du commercial (alias) transaction par transaction.
    Pour chaque ligne où From name == alias ou To name == alias :
      - la contrepartie est l'autre nom
      - le msisdn de la contrepartie est pris dans From msisdn / To msisdn si disponible

    Retourne une liste de dicts groupés par (date_op, msisdn_contrepartie) :
    {date_op, nom_contrepartie, msisdn_contrepartie, nb_transactions}

    Si les colonnes msisdn sont absentes, msisdn_contrepartie = nom_contrepartie
    (fallback sur le nom pour assurer la traçabilité même sans numéro).
    """
    alias_upper = alias.strip().upper()
    has_msisdn  = "From msisdn" in df.columns and "To msisdn" in df.columns

    records = []
    for _, row in df.iterrows():
        from_name = str(row["From name"]).strip().upper()
        to_name   = str(row["To name"]).strip().upper()

        if from_name == alias_upper:
            nom_cp   = row["To name"]
            msisdn_cp = str(row.get("To msisdn", "")).strip() if has_msisdn else ""
        elif to_name == alias_upper:
            nom_cp   = row["From name"]
            msisdn_cp = str(row.get("From msisdn", "")).strip() if has_msisdn else ""
        else:
            continue

        if not msisdn_cp or msisdn_cp in ("", "nan", "None"):
            msisdn_cp = nom_cp  # fallback nom si pas de msisdn

        records.append({
            "date_op":           str(row["Date"]),
            "nom_contrepartie":  nom_cp,
            "msisdn_contrepartie": msisdn_cp,
        })

    # Agréger par (date_op, msisdn_contrepartie)
    agg: dict[tuple, dict] = {}
    for r in records:
        key = (r["date_op"], r["msisdn_contrepartie"])
        if key not in agg:
            agg[key] = {
                "date_op":            r["date_op"],
                "nom_contrepartie":   r["nom_contrepartie"],
                "msisdn_contrepartie": r["msisdn_contrepartie"],
                "nb_transactions":    0,
            }
        agg[key]["nb_transactions"] += 1

    return list(agg.values())


# ---------------------------------------------------------------------------
# Extraction appro / destockage depuis le classeur généré
# ---------------------------------------------------------------------------

def extract_appro_from_workbook(wb_path, alias: str) -> list[dict]:
    """
    Lit le classeur Excel généré par build_transactions_workbook() et extrait
    les données d'appro et de destockage de l'alias du commercial.

    Logique :
      - Onglet "TCD - From Name" → lignes où From name == alias
          → pour chaque date : montant + nb_transactions = APPRO du commercial
            (le commercial reçoit de l'argent depuis son alias → approvisionnement)
      - Onglet "TCD - To Name"   → lignes où To name == alias
          → pour chaque date : montant + nb_transactions = DESTOCKAGE du commercial
            (le commercial envoie de l'argent vers son alias → destockage)

    Si l'alias n'est pas trouvé dans un onglet pour une date → zéro (pas d'erreur).

    Retourne une liste de dicts :
    [{date_op, type_op ('appro'|'destockage'), montant, nb_ops}, ...]
    """
    if isinstance(wb_path, (str, Path)):
        wb = load_workbook(wb_path, read_only=True, data_only=True)
    else:
        wb_path.seek(0)
        wb = load_workbook(BytesIO(wb_path.read()), read_only=True, data_only=True)

    alias_upper = alias.strip().upper()
    resultats = []

    for sheet_name, type_op in [("TCD - From Name", "appro"), ("TCD - To Name", "destockage")]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 5:
            continue

        # Ligne 4 (index 3) = dates (format "dd/mm/YYYY") sur colonnes paires
        # Ligne 5 (index 4) = sous-headers "Somme Montant" / "Nb Transactions"
        # Ligne 6+ (index 5+) = données
        # Col A (index 0) = nom

        header_dates = rows[3]   # ligne row1 : dates fusionnées
        header_sub   = rows[4]   # ligne row2 : Somme / Nb

        # Construire le mapping colonne → date
        # Les dates apparaissent en colonne paire (col 1, 3, 5, ...) dans row1
        # puis sont répétées None dans row2 (cellules fusionnées)
        date_cols: dict[int, str] = {}  # col_index → date_iso
        current_date = None
        for col_i, val in enumerate(header_dates):
            if col_i == 0:
                continue  # colonne nom
            if val is not None:
                # Essayer de parser la date
                d = _parse_date_header(val)
                if d:
                    current_date = d
            if current_date and header_sub[col_i] is not None:
                sous = str(header_sub[col_i]).strip().lower()
                if "somme" in sous or "montant" in sous:
                    date_cols[col_i] = ("montant", current_date)
                elif "nb" in sous or "transaction" in sous:
                    date_cols[col_i] = ("nb", current_date)

        # Chercher la ligne de l'alias
        for row in rows[5:]:
            if row[0] is None:
                continue
            nom_cell = str(row[0]).strip().upper()
            if nom_cell == alias_upper or alias_upper in nom_cell:
                # Extraire montant + nb par date
                par_date: dict[str, dict] = {}
                for col_i, (type_val, date_iso) in date_cols.items():
                    if col_i >= len(row):
                        continue
                    val = row[col_i]
                    if val is None or str(val).strip() in ("", "-", "nan"):
                        continue
                    try:
                        v = float(str(val).replace(",", "").replace(" ", ""))
                    except ValueError:
                        continue
                    if v == 0:
                        continue
                    if date_iso not in par_date:
                        par_date[date_iso] = {"montant": 0.0, "nb": 0}
                    if type_val == "montant":
                        par_date[date_iso]["montant"] = v
                    else:
                        par_date[date_iso]["nb"] = int(v)

                for date_iso, vals in par_date.items():
                    if vals["montant"] > 0 or vals["nb"] > 0:
                        resultats.append({
                            "date_op": date_iso,
                            "type_op": type_op,
                            "montant": vals["montant"],
                            "nb_ops":  vals["nb"],
                        })
                break  # alias trouvé, inutile de continuer

    wb.close()
    return resultats


def _parse_date_header(val) -> str | None:
    """Tente de parser un en-tête de date (str dd/mm/YYYY ou objet date)."""
    if val is None:
        return None
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            from datetime import datetime
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# Génération du classeur Excel
# ---------------------------------------------------------------------------

def _style_header(ws, row, ncols, fill=HEADER_FILL, size=10):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font  = Font(name=FONT_NAME, size=size, bold=True, color="FFFFFF")
        cell.fill  = PatternFill("solid", fgColor=fill)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_data_sheet(wb: Workbook, df: pd.DataFrame, source_label: str):
    ws = wb.create_sheet("Données")
    ws.sheet_view.showGridLines = False

    ws["A1"] = f"Données nettoyées — Transferts ({source_label})"
    ws["A1"].font = Font(name=FONT_NAME, size=13, bold=True, color=HEADER_FILL)
    ws["A2"] = (
        f"{len(df)} transactions (Type=Transfer, "
        f"hors ALBARKA GN SARL / ALBARKA GN SARL 5)"
    )
    ws["A2"].font = Font(name=FONT_NAME, size=10, italic=True, color=GREY)

    # Colonnes à afficher (on inclut les msisdn si présents)
    base_cols = ["Date", "Type", "From name", "To name", "Amount"]
    msisdn_cols = [c for c in ["From msisdn", "To msisdn"] if c in df.columns]
    cols = base_cols + msisdn_cols

    start_row = 4
    for j, h in enumerate(cols, start=1):
        ws.cell(row=start_row, column=j, value=h)
    _style_header(ws, start_row, len(cols))

    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(cols, start=1):
            val = row[col] if col in df.columns else ""
            if col == "Date" and hasattr(val, "strftime"):
                val = val.strftime("%d/%m/%Y")
            cell = ws.cell(row=i, column=j, value=val if pd.notna(val) else "")
            cell.font   = Font(name=FONT_NAME, size=10)
            cell.border = BORDER
            if col == "Amount":
                cell.number_format = "#,##0"
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=BAND_FILL)

    widths = {
        "Date": 12, "Type": 10, "From name": 36, "To name": 36,
        "Amount": 14, "From msisdn": 14, "To msisdn": 14,
    }
    for j, h in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(h, 14)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


def _write_pivot_sheet(wb: Workbook, sheet_name: str, title: str,
                       name_col: str, df: pd.DataFrame):
    """Tableau croisé dynamique par nom (To name ou From name) × Date."""
    dates   = sorted(df["Date"].unique())
    grouped = df.groupby([name_col, "Date"])["Amount"].agg(
        ["sum", "count"]
    ).reset_index()
    names = sorted(df[name_col].unique())

    rows_data = []
    for name in names:
        total_sum, total_cnt = 0, 0
        per_date = {}
        for d in dates:
            match = grouped[(grouped[name_col] == name) & (grouped["Date"] == d)]
            s = int(match["sum"].sum())   if len(match) else 0
            c = int(match["count"].sum()) if len(match) else 0
            per_date[d] = (s, c)
            total_sum += s
            total_cnt += c
        rows_data.append({
            "name": name, "per_date": per_date,
            "total_sum": total_sum, "total_cnt": total_cnt,
        })
    rows_data.sort(key=lambda r: r["total_sum"], reverse=True)

    ws = wb.create_sheet(sheet_name[:31])
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT_NAME, size=13, bold=True, color=HEADER_FILL)
    ws["A2"] = f"{len(rows_data)} entrées distinctes — Montants en XAF"
    ws["A2"].font = Font(name=FONT_NAME, size=10, italic=True, color=GREY)

    row1, row2 = 4, 5
    ws.cell(row=row1, column=1, value=name_col)
    ws.merge_cells(start_row=row1, start_column=1, end_row=row2, end_column=1)
    col = 2
    for d in dates:
        ws.cell(row=row1, column=col, value=d.strftime("%d/%m/%Y"))
        ws.merge_cells(start_row=row1, start_column=col, end_row=row1, end_column=col + 1)
        ws.cell(row=row2, column=col,     value="Somme Montant")
        ws.cell(row=row2, column=col + 1, value="Nb Transactions")
        col += 2
    ws.cell(row=row1, column=col, value="TOTAL")
    ws.merge_cells(start_row=row1, start_column=col, end_row=row1, end_column=col + 1)
    ws.cell(row=row2, column=col,     value="Somme Montant")
    ws.cell(row=row2, column=col + 1, value="Nb Transactions")
    total_col = col
    ncols     = col + 1

    _style_header(ws, row1, ncols)
    _style_header(ws, row2, ncols, fill=SUBHEADER_FILL, size=9)

    data_start = row2 + 1
    for i, r in enumerate(rows_data, start=data_start):
        ws.cell(row=i, column=1, value=r["name"]).font = Font(name=FONT_NAME, size=10)
        ws.cell(row=i, column=1).border = BORDER
        col = 2
        for d in dates:
            s, c = r["per_date"][d]
            cs = ws.cell(row=i, column=col,     value=s if c > 0 else "")
            cc = ws.cell(row=i, column=col + 1, value=c if c > 0 else "")
            cs.number_format = "#,##0"
            for cell in (cs, cc):
                cell.font      = Font(name=FONT_NAME, size=10)
                cell.border    = BORDER
                cell.alignment = Alignment(horizontal="center")
            col += 2
        cs = ws.cell(row=i, column=total_col,     value=r["total_sum"])
        cc = ws.cell(row=i, column=total_col + 1, value=r["total_cnt"])
        cs.number_format = "#,##0"
        for cell in (cs, cc):
            cell.font      = Font(name=FONT_NAME, size=10, bold=True)
            cell.border    = BORDER
            cell.alignment = Alignment(horizontal="center")
            cell.fill      = PatternFill("solid", fgColor=TOTAL_BG)
        if i % 2 == 0:
            for c_ in range(2, total_col):
                cell = ws.cell(row=i, column=c_)
                if not cell.fill or cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = PatternFill("solid", fgColor="F7FAFC")

    gt_row = data_start + len(rows_data)
    ws.cell(row=gt_row, column=1, value="TOTAL GÉNÉRAL")
    col = 2
    for d in dates:
        s = sum(r["per_date"][d][0] for r in rows_data)
        c = sum(r["per_date"][d][1] for r in rows_data)
        cs = ws.cell(row=gt_row, column=col,     value=s)
        cc = ws.cell(row=gt_row, column=col + 1, value=c)
        cs.number_format = "#,##0"
        col += 2
    tot_s = sum(r["total_sum"] for r in rows_data)
    tot_c = sum(r["total_cnt"] for r in rows_data)
    ws.cell(row=gt_row, column=total_col,     value=tot_s).number_format = "#,##0"
    ws.cell(row=gt_row, column=total_col + 1, value=tot_c)
    _style_header(ws, gt_row, ncols)

    ws.column_dimensions["A"].width = 42
    for c_ in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(c_)].width = 15
    ws.freeze_panes = ws.cell(row=data_start, column=2)


def build_transactions_workbook(df: pd.DataFrame, source_label: str = "") -> Workbook:
    """
    Construit le classeur complet :
      - Onglet "Données"       : transactions nettoyées (avec msisdn si présents)
      - Onglet "TCD - To Name" : tableau croisé par bénéficiaire × date
      - Onglet "TCD - From Name" : tableau croisé par expéditeur × date
    """
    wb = Workbook()
    wb.remove(wb.active)
    _write_data_sheet(wb, df, source_label)
    _write_pivot_sheet(wb, "TCD - To Name",
                       "Tableau croisé — Par bénéficiaire (To Name)", "To name", df)
    _write_pivot_sheet(wb, "TCD - From Name",
                       "Tableau croisé — Par expéditeur (From Name)", "From name", df)
    return wb
